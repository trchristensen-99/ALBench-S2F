"""Sequence-budget workbook -- live-editable on Zoom, Slides-sized paste block.

LAYOUT.  A:J rows 8..31 are the paste block (24 x 10 at 7 pt = 8.6 x 4.50 in, fits 16:9). Toggles sit
in L:R on the SAME ROWS, so inserting/deleting a strategy updates table and controls in one action and
the numbers move next to the switch you flip. Blank toggles are INERT -- an arm counts only if its
cell holds exactly 1 (or >1 to pin) -- otherwise an inserted blank row would silently add nA arms.

The per-row total column was dropped from the paste block on purpose: in a margins design every
single-reservoir row has exactly ONE funded cell, so its row total is trivially the per-arm size.
Row totals live in column S, outside the paste block. That freed the 10th column for Rafi's second
diversity method.

Allocation is EVEN across active arms; synthesis-cost differences are handled by adjusting the
scaling SLOPE per unit cost afterwards, never by ordering unequal numbers of sequences.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.expanduser("~/Downloads/joint_PI_meeting_aug20/sequence_budget.xlsx")

# (family, reservoir, cell-type role, tier)   tier A = single-R arm, B = main pool, C = sub-pool
RESERVOIRS = [
    ("Random", "Uniform ACGT", "agn", "A"),
    ("Random", "Dinuc-shuffled genomic", "agn", "A"),
    ("Random", "Composition-tuned (GC / k-mer biased)", "agn", "A"),
    ("Genomic", "ENCODE accessible - 4 accessibility strata", "diff", "A"),
    ("Genomic", "Zoonomia ortholog CREs (phylogenetic)", "agn", "A"),
    ("Genomic", "Gosai alts - TRAIN chr only (test chr held out)", "agn", "A"),
    ("Genomic pert.", "Mutagenesis - oracle-tuned rate (SNV pairs)", "agn", "A"),
    ("Genomic pert.", "Structural - indel/transloc/inv, oracle-tuned", "agn", "A"),
    ("Genomic pert.", "EvoAug - as-published ML-aug settings (ctrl)", "agn", "A"),
    ("Motif-based", "Coverage - FULL vocabulary (~600 motifs)", "agn", "A"),
    ("Motif-based", "Coverage - CT vocabulary, 4 strata", "diff", "A"),
    ("Motif-based", "Syntax core - RESTRICTED vocab (~60), crossed", "agn", "A"),
    ("Model-gen.", "DNA-LM (HyenaDNA) - unconditioned", "agn", "A"),
    ("Model-gen.", "D3 diffusion - activity-conditioned", "diff", "A"),
    ("Model-gen.", "D3 diffusion - genomic-conditioned", "agn", "A"),
    ("Model-gen.", "D3 diffusion - uncertainty-conditioned", "diff", "A"),
    ("Model-gen.", "In-silico directed evolution (oracle-guided)", "diff", "A"),
    ("POOLED", "All reservoirs, equal parts", "agn", "B"),
    ("POOLED", "Genomic-derived only", "agn", "C"),
    ("POOLED", "Synthetic / generated only", "agn", "C"),
]
ACQ = [
    "Random",
    "Uncertainty (joint)",
    "Uncertainty (per CT)",
    "Diversity (model latent)",
    "Diversity (TFBS embed.)",
    "Uncert. x Divers.",
    "Activity-strat. (2D joint)",
]

ON = set()
for i, (_f, _n, _ct, tier) in enumerate(RESERVOIRS):
    if tier == "A":
        ON.add((i, 0))
    if tier == "B":
        ON.update((i, j) for j in range(len(ACQ)))
    if tier == "C":
        ON.update({(i, 1), (i, 3)})

HDR = PatternFill("solid", fgColor="1E293B")
CTRLH = PatternFill("solid", fgColor="92400E")
EDIT = PatternFill("solid", fgColor="FEF3C7")
OFFF = PatternFill("solid", fgColor="F1F5F9")
CALC = PatternFill("solid", fgColor="F8FAFC")
SUMF = PatternFill("solid", fgColor="DBEAFE")
POOL = PatternFill("solid", fgColor="FDE68A")
FAMF = {
    "Random": "EDE9FE",
    "Genomic": "DBEAFE",
    "Genomic pert.": "CCFBF1",
    "Motif-based": "FEF9C3",
    "Model-gen.": "FCE7F3",
    "POOLED": "FDE68A",
}
CTF = {
    "joint": PatternFill("solid", fgColor="E0E7FF"),
    "diff": PatternFill("solid", fgColor="FEE2E2"),
    "agn": None,
}
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

nR, nA = len(RESERVOIRS), len(ACQ)
POOL_I = [i for i, r in enumerate(RESERVOIRS) if r[3] != "A"]
C0 = 4
LASTC = C0 + nA - 1
AH, A0 = 8, 9
ALAST, SROW = A0 + nR - 1, A0 + nR
K0 = LASTC + 2
TOTC = K0 + nA
LC = get_column_letter(LASTC)
KC0, KCL = get_column_letter(K0), get_column_letter(K0 + nA - 1)
CRNG = f"${KC0}${A0}:${KCL}${ALAST}"
PRNG = f"${KC0}${A0 + POOL_I[0]}:${KCL}${A0 + POOL_I[-1]}"

wb = Workbook()
ws = wb.active
ws.title = "Budget"
ws["A1"] = "Sequence budget - funded arms"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = (
    f"Paste A{AH}:{LC}{SROW} into Slides ({SROW - AH + 1} x {nA + 3} at 8 pt). "
    f"Toggles in {KC0}:{KCL} on the SAME ROWS: 1 = on, 0/blank = off, >1 = pin that count."
)
ws["A2"].font = Font(italic=True, size=9, color="666666")


def lbl(cell, t, **kw):
    ws[cell] = t
    ws[cell].font = Font(size=9, **kw)


lbl("A3", "Synthesis budget", bold=True)
ws["C3"] = 3_200_000
ws["C3"].fill = EDIT
ws["C3"].number_format = "#,##0"
ws["C3"].border = BOX
ws["C3"].font = Font(bold=True, size=9)
lbl("D3", "sequences ORDERED (pooled arms share them)", italic=True, color="666666")
lbl("A4", "Pooled-arm overlap", bold=True)
ws["C4"] = 0.0
ws["C4"].fill = EDIT
ws["C4"].number_format = "0%"
ws["C4"].border = BOX
ws["C4"].font = Font(bold=True, size=9)
lbl(
    "D4",
    "mean pairwise Jaccard of selections - MEASURE IN SILICO before ordering",
    italic=True,
    color="B45309",
)

n_pool = f"COUNTIF({PRNG},1)"
n_all = f"COUNTIF({CRNG},1)"
n_arms = f'({n_all}+COUNTIF({CRNG},">1"))'
eff = f"({n_all}-{n_pool}+IF({n_pool}=0,0,1+MAX(0,{n_pool}-1)*(1-$C$4)))"

lbl("A5", "Arms funded", bold=True)
ws["C5"] = f"={n_arms}"
ws["C5"].font = Font(bold=True, size=9)
lbl("D5", "Effective cells", bold=True)
ws["F5"] = f"=ROUND({eff},1)"
ws["F5"].font = Font(bold=True, size=9)
lbl("G5", "Sequences / arm", bold=True)
ws["I5"] = f'=ROUND(($C$3-SUMIF({CRNG},">1"))/MAX(1,{eff}),-2)'
ws["I5"].number_format = "#,##0"
ws["I5"].font = Font(bold=True, size=11)
ws["I5"].fill = SUMF
ws["I5"].border = BOX
ws["A6"] = (
    '=IF(I5>=100000,"OK "&TEXT(I5,"#,##0")&"/arm - from-scratch curve 1k to "&TEXT(I5,"#,##0")'
    '&" (~2 OOMs), 3 replicates at each of the 3 smallest points",'
    'IF(I5>=30000,"1k to 30k only (~1.5 OOMs) - prune arms to recover the curve",'
    '"TOO SMALL for a per-arm curve"))'
)
ws["A6"].font = Font(bold=True, size=9, color="15803D")


def header(row, fill, first, size=7.0):
    for c, t in ((1, "Family"), (2, first), (3, "CT")):
        h = ws.cell(row, c, t)
        h.fill = fill
        h.font = Font(bold=True, color="FFFFFF", size=size + 0.5)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, a in enumerate(ACQ):
        h = ws.cell(row, C0 + j, a)
        h.fill = fill
        h.font = Font(bold=True, color="FFFFFF", size=size)
        h.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


header(AH, HDR, "Reservoir strategy")

for i, (fam, name, ct, tier) in enumerate(RESERVOIRS):
    r = A0 + i
    f = ws.cell(r, 1, fam if (i == 0 or RESERVOIRS[i - 1][0] != fam) else "")
    f.font = Font(size=7, bold=True)
    f.fill = PatternFill("solid", fgColor=FAMF[fam])
    f.alignment = Alignment(vertical="center", wrap_text=True)
    n = ws.cell(r, 2, name)
    n.font = Font(size=8, bold=tier != "A")
    n.alignment = Alignment(vertical="center")
    if tier != "A":
        n.fill = POOL
    c = ws.cell(r, 3, ct)
    c.font = Font(size=6.5)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if CTF[ct]:
        c.fill = CTF[ct]
    for j in range(nA):
        kcol = get_column_letter(K0 + j)
        cell = ws.cell(r, C0 + j, f'=IF({kcol}{r}=1,$I$5,IF({kcol}{r}>1,{kcol}{r},"-"))')
        cell.number_format = "#,##0"
        cell.border = BOX
        cell.font = Font(size=8)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if (i, j) not in ON:
            cell.fill = OFFF
        elif i % 2:
            cell.fill = CALC
    t = ws.cell(r, TOTC, f'=SUMIF({get_column_letter(C0)}{r}:{LC}{r},">0")')
    t.number_format = "#,##0"
    t.fill = SUMF
    t.border = BOX
    t.font = Font(size=8, bold=True)

ws.cell(SROW, 2, "Sum of arm sizes").font = Font(size=7.5, bold=True)
for j in range(nA):
    col = get_column_letter(C0 + j)
    c = ws.cell(SROW, C0 + j, f'=SUMIF({col}{A0}:{col}{ALAST},">0")')
    c.number_format = "#,##0"
    c.fill = SUMF
    c.border = BOX
    c.font = Font(size=7.5, bold=True)
    c.alignment = Alignment(horizontal="right")
g = ws.cell(SROW, TOTC, f"=SUM({get_column_letter(C0)}{SROW}:{LC}{SROW})")
g.number_format = "#,##0"
g.font = Font(size=8, bold=True)
g.border = BOX
g.fill = PatternFill("solid", fgColor="BBF7D0")
ws.cell(
    SROW + 1,
    1,
    f'="ordered "&TEXT(ROUND($I$5*{eff},-2),"#,##0")&"   |   "&'
    f'IF(ABS(ROUND($I$5*{eff},-2)-$C$3)<20000,"within budget","MISMATCH")&'
    '"   |   CT: agn = agnostic, joint = shared/both, diff = differential"',
).font = Font(bold=True, size=8)

ws.cell(AH - 1, K0, "CONTROL - 1 = on, 0/blank = off, >1 = pin").font = Font(
    bold=True, size=9, color="92400E"
)
for j, a in enumerate(ACQ):
    h = ws.cell(AH, K0 + j, a.split(" (")[0][:10])
    h.fill = CTRLH
    h.font = Font(bold=True, color="FFFFFF", size=6.5)
    h.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
h = ws.cell(AH, TOTC, "row sum")
h.fill = HDR
h.font = Font(bold=True, color="FFFFFF", size=6.5)
h.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
for i in range(nR):
    for j in range(nA):
        cell = ws.cell(A0 + i, K0 + j, 1 if (i, j) in ON else 0)
        cell.fill = EDIT if (i, j) in ON else OFFF
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=8)

WA, WB, WC, WD = 10.5, 30.0, 4.5, 9.6
for col, w in (("A", WA), ("B", WB), ("C", WC)):
    ws.column_dimensions[col].width = w
for j in range(nA):
    ws.column_dimensions[get_column_letter(C0 + j)].width = WD
ws.column_dimensions[get_column_letter(LASTC + 1)].width = 2.0
for j in range(nA):
    ws.column_dimensions[get_column_letter(K0 + j)].width = 7.5
ws.column_dimensions[get_column_letter(TOTC)].width = 9.5
for r in range(AH - 1, SROW + 1):
    ws.row_dimensions[r].height = 14.0
ws.freeze_panes = "D9"

# ==================== FOR THE MEETING ====================
wsf = wb.create_sheet("ForMeeting")
wsf["A1"] = "Open decisions - bring these to the PIs"
wsf["A1"].font = Font(bold=True, size=13)
DEC = [
    (
        "Gosai alts as a reservoir?",
        "Rafi: it is our test data.",
        "Only the ~35k chromosome-held-out alts are test. The ~300k train-chromosome alts are usable "
        "with no leakage, and we already own their Gosai measurements - so re-measure only a subset for "
        "assay calibration. KEEP, relabelled 'TRAIN chr only'. Decide: full arm, or calibration subset "
        "folded into the anchor set?",
    ),
    (
        "Drop 'open in BOTH' and 'shared-core motifs'?",
        "Rafi: overkill if we already have per-cell-type accessible and TFBS-rich arms - we will see "
        "informativeness from those.",
        "Counter: the shared arms are the only ones that test whether SHARED regulatory information is "
        "sufficient, which is the null for the whole cell-type-specificity story. Dropping them saves "
        "2 arms; keeping them costs ~9k/arm. Middle option: merge shared + differential into ONE arm "
        "with a within-arm split, keeping the contrast at half the cost.",
    ),
    (
        "Drop 'Uncertainty (per CT)'?",
        "Rafi: may be able to skip it.",
        "It is the only test of H6 (does per-cell-type selection beat joint?). Costs 1 arm because it "
        "is N/2+N/2, not 2N. Cheap way to retire a question that will otherwise be asked at review.",
    ),
    (
        "Two diversity methods - ADDED per Rafi",
        "Model-latent diversity vs TFBS-embedding diversity (so all TFs are represented).",
        "Worth keeping both: one covers what the MODEL represents, the other covers TF space regardless "
        "of the model. When they disagree, that tells you whether informativeness is model-defined or "
        "biology-defined. Both are cell-type agnostic, so neither needs a per-CT variant.",
    ),
    (
        "Motif syntax vs composition - CLARIFIED",
        "Rafi asked what separates 'motif grammar' from 'TFBS recombination'.",
        "Two axes of one generator. COMPOSITION = which motifs are present (identity and count) - "
        "recombining real TFBS instances into backgrounds. SYNTAX = hold the motif set FIXED and vary "
        "order, spacing, orientation, copy number, distance to the reporter. Composition asks which TFs "
        "matter; syntax asks whether arrangement matters. Now separate rows.",
    ),
    (
        "EvoAug as the H5 control - REFRAMED",
        "Its published settings were tuned for ML data AUGMENTATION (regularisation during training), "
        "not for generating sequences to be measured.",
        "That makes it a BETTER control, not a worse one: it is exactly the off-the-shelf setting a "
        "practitioner would naively reuse. Keep it as the naive baseline and oracle-tune the two "
        "perturbation arms. If the PIs prefer, drop this row and test H5 inside the mutagenesis arm "
        "instead by ordering two rates (oracle-optimal and naive default).",
    ),
    (
        "D3 uncertainty-conditioned is CT-specific - FIXED",
        "Uncertainty can be defined on K562, on HepG2, or jointly.",
        "Retagged 'diff'. Default conditioning target is JOINT uncertainty, with per-cell-type strata "
        "inside the arm, so the CT question is answered without a second arm.",
    ),
    (
        "In-silico directed evolution - ADDED",
        "Oracle-guided iterative mutate-and-select.",
        "Distinct from D3 (no generative model) and from EP-PCR (random perturbation, no selection). "
        "It is the strongest activity-maximising design method and can be evolved toward the "
        "DIFFERENTIAL objective, which is Carl's gene-therapy angle. Tagged 'diff'.",
    ),
    (
        "Three random arms - ADDED",
        "Uniform ACGT, dinuc-shuffled, composition-tuned (GC / k-mer biased).",
        "These are genuinely different nulls: uniform = pure null; dinuc-shuffle = composition-matched "
        "null; composition-tuned = deliberately biased toward accessible-region composition. Together "
        "they answer 'how much of informativeness is just nucleotide composition?', which is the "
        "cheapest interpretable result in the study.",
    ),
    (
        "Row budget",
        "22 strategies x 7 acquisitions, 30 arms.",
        "Fits a slide at 7 pt. Pruning to ~18 strategies restores 8 pt and pushes each arm from ~107k "
        "to ~128k. Per-arm size is the thing to protect: below 100k the from-scratch curve loses its "
        "top decade.",
    ),
]
for j, t in enumerate(["Question", "Argument raised", "Our read / recommendation"], 1):
    c = wsf.cell(3, j, t)
    c.fill = HDR
    c.font = Font(bold=True, color="FFFFFF", size=9)
for i, row in enumerate(DEC, start=4):
    for j, v in enumerate(row, 1):
        c = wsf.cell(i, j, v)
        c.font = Font(size=8.5, bold=(j == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
    wsf.cell(i, 1).fill = PatternFill("solid", fgColor="F1F5F9")
for w, col in ((36, "A"), (48, "B"), (82, "C")):
    wsf.column_dimensions[col].width = w

# ==================== MENU ====================
wsm = wb.create_sheet("Menu")
wsm["A1"] = "Reservoir candidate menu - prune here, then set the toggles"
wsm["A1"].font = Font(bold=True, size=13)
MENU = [
    ("Random", "Uniform ACGT", "FUNDED", "experiment", "Pure null."),
    (
        "Random",
        "Dinuc-shuffled genomic",
        "FUNDED",
        "experiment",
        "Composition-matched null: preserves mono- and dinucleotide content, destroys grammar.",
    ),
    (
        "Random",
        "Composition-tuned (GC / k-mer biased)",
        "FUNDED",
        "in-silico (bias target)",
        "Bias composition toward accessible-region statistics. Separates 'composition matters' from "
        "'grammar matters' - the bias target is chosen in silico.",
    ),
    (
        "Genomic",
        "Gosai alts - TRAIN chr only",
        "FUNDED (was flagged)",
        "experiment",
        "Test chr (~35k) held out; ~300k train-chr alts are leakage-free. Measurements already exist, "
        "so a subset may suffice for assay calibration. Delta-supervision: value is the paired contrast.",
    ),
    (
        "Genomic",
        "ENCODE accessible - open in BOTH",
        "PI DECISION",
        "experiment",
        "Rafi: possibly redundant with the per-CT arms. Counter: the only test of whether SHARED "
        "regulatory information suffices.",
    ),
    (
        "Genomic",
        "ENCODE accessible - CT-differential",
        "FUNDED",
        "experiment",
        "K562-only and HepG2-only in ONE arm, factorial within it, assayed in both cell types.",
    ),
    (
        "Genomic",
        "Zoonomia ortholog CREs",
        "FUNDED",
        "in-silico (clade cutoff)",
        "Evolutionary constraint as the prior; the only cross-species reservoir. OPEN: clade cutoff - "
        "ask Anirban.",
    ),
    (
        "Genomic",
        "Closed / inactive genomic regions",
        "PROPOSED",
        "experiment",
        "True negatives from real sequence. Random and dinuc-shuffle give synthetic low activity; "
        "closed chromatin is a different distribution and may be needed to calibrate the low end.",
    ),
    (
        "Genomic pert.",
        "Mutagenesis - oracle-tuned rate",
        "FUNDED",
        "in-silico",
        "Screen the rate on the oracle, order ONE tuned rate.",
    ),
    (
        "Genomic pert.",
        "Structural - indel/transloc/inv",
        "FUNDED (merged)",
        "in-silico",
        "Three separate arms would spend budget answering a mix question the oracle can screen.",
    ),
    (
        "Genomic pert.",
        "EvoAug - as-published settings",
        "FUNDED as H5 control",
        "experiment",
        "Published values were tuned for ML augmentation, NOT for generating measured data - which is "
        "precisely why it is the naive off-the-shelf baseline.",
    ),
    (
        "Motif-based",
        "Composition - full motif DB",
        "FUNDED",
        "in-silico (params)",
        "The genuinely CT-agnostic motif arm: whole known DB, NO expression filtering. Teaches grammar, "
        "not cell-type identity.",
    ),
    (
        "Motif-based",
        "Composition - shared-core motifs",
        "PI DECISION",
        "in-silico (params)",
        "TFs expressed in both lines (MYC / AP1 core). Rafi: possibly redundant.",
    ),
    (
        "Motif-based",
        "Composition - CT-enriched mixed",
        "FUNDED",
        "in-silico (params)",
        "Within-arm 2x2: K562-enriched / HepG2-enriched / BOTH in one sequence / neither. The BOTH cell "
        "teaches differential activity, attribution intact.",
    ),
    (
        "Motif-based",
        "Syntax - order/spacing/orientation/copies",
        "FUNDED",
        "in-silico (params)",
        "Motif set held FIXED, arrangement varied. This is the axis distinct from composition.",
    ),
    (
        "Model-gen.",
        "DNA-LM (HyenaDNA) - unconditioned",
        "FUNDED",
        "experiment",
        "Generative prior, no activity signal.",
    ),
    (
        "Model-gen.",
        "D3 diffusion - unconditioned",
        "PRUNE",
        "in-silico",
        "Redundant with HyenaDNA-unconditioned; conditioning is the interesting axis.",
    ),
    (
        "Model-gen.",
        "D3 diffusion - activity-conditioned",
        "FUNDED",
        "experiment",
        "Carl: conditional generation will likely dominate by a large margin. This is that test.",
    ),
    (
        "Model-gen.",
        "D3 diffusion - genomic-conditioned",
        "FUNDED",
        "experiment",
        "Genomic realism without activity targeting.",
    ),
    (
        "Model-gen.",
        "D3 diffusion - uncertainty-conditioned",
        "FUNDED (CT-tagged)",
        "experiment",
        "GENERATION as active learning. Conditioned on JOINT uncertainty with per-CT strata in-arm.",
    ),
    (
        "Model-gen.",
        "In-silico directed evolution",
        "PROPOSED - ADDED",
        "in-silico (rounds)",
        "Oracle-guided mutate-and-select. Distinct from diffusion (no generative model) and from EP-PCR "
        "(no selection). Can target activity OR the K562-HepG2 differential.",
    ),
    (
        "Model-gen.",
        "Gradient-based adversarial sequences",
        "PROPOSED",
        "in-silico",
        "Direct ascent on model uncertainty/disagreement. Overlaps uncertainty-conditioned D3; include "
        "only if the two are shown to differ in silico.",
    ),
    (
        "Other",
        "Public MPRA sequences from other cell lines",
        "PROPOSED",
        "experiment",
        "Re-measuring existing libraries in our two lines tests transfer across studies and gives a "
        "cross-dataset calibration anchor. Cheap: sequences already designed.",
    ),
]
for j, t in enumerate(["Family", "Variant", "Status", "Decide by", "Rationale"], 1):
    c = wsm.cell(3, j, t)
    c.fill = HDR
    c.font = Font(bold=True, color="FFFFFF", size=9)
for i, row in enumerate(MENU, start=4):
    for j, v in enumerate(row, 1):
        c = wsm.cell(i, j, v)
        c.font = Font(size=8.5, bold=(j == 3))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if j == 1:
            c.fill = PatternFill("solid", fgColor=FAMF.get(row[0], "FFFFFF"))
for w, col in ((14, "A"), (34, "B"), (22, "C"), (18, "D"), (76, "E")):
    wsm.column_dimensions[col].width = w

# ==================== DECISIONS ====================
wsd = wb.create_sheet("Decisions")
wsd["A1"] = "Every design choice, and why"
wsd["A1"].font = Font(bold=True, size=13)
DD = [
    (
        "BUDGET",
        "Even N across funded arms; weights removed",
        "Unequal N confounds synthesis cost with informativeness. Cost enters later as a per-unit-cost "
        "adjustment to the scaling slope, testing the same N per strategy.",
    ),
    (
        "BUDGET",
        "Order the UNION; overlap is a discount, not a surcharge",
        "Pooled arms picking the same sequence synthesise it once, so overlap BUYS larger arms. "
        "Budget never expands.",
    ),
    (
        "GRID",
        "Fund the MARGINS (a plus-shape), not the interior",
        "ANOVA main effects need the margins. The full 22x7 grid is 154 cells at ~21k each - one point "
        "per cell, no replicates, no curve. 30 arms at ~107k each buys curves WITH error bars.",
    ),
    (
        "GRID",
        "Per-row total dropped from the paste block",
        "In a margins design each single-reservoir row has exactly ONE funded cell, so its row total is "
        "trivially the per-arm size. Freed the 10th column for the second diversity method.",
    ),
    (
        "R EFFECT",
        "Single-reservoir arms run at acquisition = random",
        "Peter: reservoirs are just sampling. Random keeps the reservoir measurement unconfounded.",
    ),
    (
        "A EFFECT",
        "Pooled arms: all acquisitions select from ONE union pool",
        "Pool held fixed means acquisition is the only varying factor.",
    ),
    (
        "R x A",
        "Interaction recovered as composition, not funded as cells",
        "Enrichment = picked/expected in the pooled arms. Free, but preference is NOT value.",
    ),
    (
        "ANALYSIS",
        "Three curves from the SAME ordered sequences",
        "A1 from-scratch per-arm (~2 OOMs) | A2 marginal over the 300k base (~0.5 OOM) | A3 union "
        "corpus 10k-3.5M (~2.5 OOMs, the headline).",
    ),
    (
        "ANALYSIS",
        "Level is confirmatory, slope is exploratory",
        "Our slope-variance analysis found informativeness shows up as LEVEL not RATE; slopes were "
        "indistinguishable at 3 D points. Having the prior null makes this defensible.",
    ),
    (
        "REPLICATES",
        "3 models on DISJOINT subsets at each of the 3 smallest D",
        "Carl: one sequence set measures THAT DATASET, not the method. Costs no extra sequences.",
    ),
    ("REPLICATES", "Nest 10k -> 30k -> N, never resample from scratch", "Rafi."),
    (
        "RANDOM",
        "Three separate random arms",
        "Uniform = pure null; dinuc-shuffle = composition-matched null; composition-tuned = biased "
        "toward accessible-region statistics. Together they isolate how much of informativeness is "
        "just nucleotide composition.",
    ),
    (
        "MOTIF",
        "COMPOSITION and SYNTAX are separate axes",
        "Composition = which motifs are present (recombining TFBS instances). Syntax = motif set fixed, "
        "vary order/spacing/orientation/copy number. Composition asks which TFs matter; syntax asks "
        "whether arrangement matters.",
    ),
    (
        "MOTIF",
        "Motif provenance has 3 levels; 'both' is NOT a union",
        "Full DB unfiltered (agnostic) | expressed in BOTH | CT-enriched. Only the unfiltered level "
        "isolates grammar from cell-type identity. Mixing lives INSIDE the CT-enriched arm as a 2x2.",
    ),
    (
        "MOTIF",
        "Count, order, spacing, background decided IN SILICO",
        "Enumerating them experimentally would cost ~10 arms to answer what the oracle can screen.",
    ),
    (
        "PERTURB",
        "Indel + translocation + inversion folded into ONE tuned arm",
        "Three arms would spend budget on a mix question the oracle can screen.",
    ),
    (
        "PERTURB",
        "EvoAug as-published is the H5 naive control",
        "Its settings were tuned for ML augmentation, not for generating measured data - exactly the "
        "off-the-shelf choice a practitioner would reuse. Without a control, in-silico tuning is an "
        "untested assumption.",
    ),
    (
        "MODEL-GEN",
        "In-silico directed evolution added",
        "Oracle-guided mutate-and-select: no generative model (unlike D3), with selection (unlike "
        "EP-PCR). Can target activity or the K562-HepG2 differential.",
    ),
    (
        "MODEL-GEN",
        "Uncertainty-conditioned D3 tagged CT-differential",
        "Uncertainty is definable on K562, HepG2 or jointly. Default is joint with per-CT strata "
        "in-arm, so no second arm is needed.",
    ),
    (
        "ACQ",
        "Two diversity methods (Rafi)",
        "Model-latent covers what the MODEL represents; TFBS-embedding covers TF space regardless of "
        "the model. Disagreement between them says whether informativeness is model- or "
        "biology-defined. Both CT-agnostic, so neither needs a per-CT variant.",
    ),
    (
        "ACQ",
        "Per-CT arm gets N/2 + N/2 = N, never 2N",
        "Matched N is required for the main-effect test; at 2N the extra data alone would win.",
    ),
    (
        "ACQ",
        "Joint-vs-per-CT tested in the uncertainty family ONLY (H6)",
        "Crossing {joint, per-CT} with every family is unaffordable. Uncertainty is per-model "
        "per-task, so unambiguously CT-dependent.",
    ),
    (
        "ACQ",
        "Contrast |K562-HepG2| folded into activity-stratified 2D",
        "Off-diagonal bins of the (K562, HepG2) prediction grid ARE contrast selections. Delivers "
        "Peter's even spread over dynamic range and Carl's specificity angle in one arm.",
    ),
    (
        "LEAKAGE",
        "Include ALL known motifs; hold out COMBINATIONS, not identities",
        "A model that never saw a motif cannot generalise to it, and including everything known is "
        "what a practitioner would do. Real risk is memorising the planting procedure. Three-way "
        "chromosome split: Gosai base is genomic and EP-PCR / EvoAug / Zoonomia are genome-seeded.",
    ),
    (
        "EVAL",
        "Battery of test sets, never a single scalar",
        "Peter: different libraries may be optimal for different purposes.",
    ),
    (
        "ASSAY",
        "Re-measure test AND val in full; ~10k train anchors",
        "Val in old units with test in new units means selecting in one space and scoring in another. "
        "Anchor strata picked from ONE HALF of Gosai's replicates, calibrated on the other half.",
    ),
    (
        "ASSAY",
        "Integrate old + new via a per-assay head on a shared trunk",
        "Handles sequence-dependent batch effects. Fit a global transform first as a diagnostic only.",
    ),
]
for j, t in enumerate(["Area", "Decision", "Why"], 1):
    c = wsd.cell(3, j, t)
    c.fill = HDR
    c.font = Font(bold=True, color="FFFFFF", size=9)
for i, (a, d, w) in enumerate(DD, start=4):
    for j, v in enumerate((a, d, w), 1):
        c = wsd.cell(i, j, v)
        c.font = Font(size=8.5, bold=(j <= 2))
        c.alignment = Alignment(wrap_text=True, vertical="top")
    wsd.cell(i, 1).fill = PatternFill("solid", fgColor="F1F5F9")
for w, col in ((13, "A"), (46, "B"), (86, "C")):
    wsd.column_dimensions[col].width = w

# ==================== HOW TO EDIT ====================
wsh = wb.create_sheet("HowToEdit")
HOW = [
    "PRESENTING ON ZOOM",
    "  Share the Excel WINDOW, not the whole screen; zoom to ~140% so the block fills the frame.",
    "  Toggles are in L:R on the same rows as the table, so the audience sees numbers move next to the",
    "  switch you flip. Keep a static pasted copy on a backup slide in case live editing gets messy.",
    "  For PIs to edit afterwards, upload to Google Sheets - all formulas are plain COUNTIF/SUMIF/IF.",
    "",
    "ADD A STRATEGY (row)",
    "  1. Right-click a row INSIDE rows 10-30 and Insert. Never at the very first or last row of the",
    "     block - Excel only auto-expands ranges for interior inserts.",
    "  2. Copy the row above and paste into the new row. This brings the D:J formulas AND the L:R",
    "     toggles across in one action, because they share rows.",
    "  3. Retype family, name and CT tag in A:C.",
    "  Blank toggles are INERT, so a new row contributes nothing until you type 1 into L:R.",
    "",
    "REMOVE A STRATEGY",
    "  Preferred: set its toggles to 0. Budget redistributes, nothing breaks, and the row stays visible",
    "  as an explicit 'considered and rejected' - useful in front of the PIs.",
    "",
    "ADD AN ACQUISITION (column)",
    "  1. Insert a column inside D:J.  2. Insert a matching column inside L:R.",
    "  3. Copy an adjacent table cell across the new column, and an adjacent toggle across the new one.",
    "  4. Type the header. Sums and the per-arm share auto-expand.",
    "  NOTE: 10 columns is the Slides ceiling. An 8th acquisition means dropping the CT column.",
    "",
    "WHAT RECALCULATES AUTOMATICALLY",
    "  I5 sequences/arm - the single driver: (budget - pinned) / effective cells.",
    "  F5 effective cells - applies the pooled-overlap discount.",
    "  Every table cell reads ITS OWN row's toggle, so nothing is hard-coded per row.",
    "  Row sums (S), column sums (row 31), ordered-vs-budget check (row 32).",
    "",
    "WHAT WILL BREAK IT",
    "  Typing a number into the TABLE (D:J) - that overwrites the formula. Type into L:R instead.",
    "  Inserting at the exact top or bottom edge of the block, leaving ranges unexpanded.",
    "  Sorting only part of the block - the toggles ride with the rows, so sort A:S together or not at all.",
]
for k, t in enumerate(HOW, start=2):
    c = wsh.cell(k, 1, t)
    c.font = Font(size=9, bold=(t.strip() != "" and not t.startswith("  ")))
wsh.column_dimensions["A"].width = 104

# ==================== SCHEDULE ====================
ws3 = wb.create_sheet("Schedule")
ws3["A1"] = "Three analyses from the same ordered sequences"
ws3["A1"].font = Font(bold=True, size=13)
rows = [
    ("", "Init", "Training-set schedule", "Span", "Answers", "Tier"),
    (
        "A1  From-scratch per-arm",
        "random",
        "nested 1k, 3k, 10k, 30k, N within each arm",
        "~2 OOMs",
        "which strategy scales better, unconfounded by the 300k base",
        "CONFIRMATORY (level)",
    ),
    (
        "A2  Marginal over base",
        "300k Gosai",
        "base + nested 10k, 30k, N",
        "~0.5 OOM",
        "we already have Gosai - what do we order next?",
        "CONFIRMATORY",
    ),
    (
        "A3  Union corpus",
        "random",
        "nested 10k ... 300k ... ~3.5M over Gosai + all ordered",
        "~2.5 OOMs",
        "how far does this go; the extrapolation for the funding case",
        "HEADLINE",
    ),
]
for i, r in enumerate(rows, start=3):
    for j, v in enumerate(r, start=1):
        c = ws3.cell(i, j, v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 3:
            c.fill = HDR
            c.font = Font(bold=True, color="FFFFFF", size=9)
        else:
            c.font = Font(size=9, bold=(j == 1))
for w, col in ((26, "A"), (11, "B"), (40, "C"), (10, "D"), (44, "E"), (21, "F")):
    ws3.column_dimensions[col].width = w
for k, t in enumerate(
    [
        "",
        "REPLICATES. At each of the 3 smallest D points, train 3 models on DISJOINT subsets of the arm.",
        "Carl: a model trained on one sequence set measures THAT DATASET; to measure the STRATEGY you",
        "must train on different sequence sets drawn from it. Nesting (Rafi) and replication (Carl) both",
        "come out of the same ordered sequences.",
        "",
        "LEVEL IS CONFIRMATORY, SLOPE IS NOT. Our slope-variance analysis found reservoir informativeness",
        "shows up as LEVEL, not RATE - slopes were indistinguishable at 3 D points. A1 gives 5 points with",
        "replicates, making slope testable, but it stays pre-registered as EXPLORATORY.",
    ],
    start=8,
):
    ws3.cell(k, 1, t).font = Font(size=9, italic=True, color="444444")

# ==================== HYPOTHESES ====================
ws2 = wb.create_sheet("Hypotheses")
H = [
    "CONFIRMATORY - powered, fixed before ordering",
    "  H1  RESERVOIR main effect. From-scratch (A1), matched N, acquisition = random. Reservoirs differ",
    "      in the LEVEL of the scaling curve.",
    "  H2  ACQUISITION main effect. Pooled arms, matched N: every acquisition selects from the SAME",
    "      union pool, so acquisition is the only thing varying.",
    "  H3  MARGINAL informativeness over the 300k Gosai base (A2) differs by reservoir.",
    "  H4  H1 and H3 RANK STRATEGIES DIFFERENTLY, and the divergence tracks distance from the existing",
    "      corpus. Pre-registering this turns the corpus-overlap confound into a result.",
    "  H5  IN-SILICO TUNING TRANSFERS. Oracle-tuned mutagenesis and structural arms beat the",
    "      as-published EvoAug control - the arm that justifies deciding parameters in silico.",
    "  H6  JOINT vs PER-CELL-TYPE acquisition, uncertainty family only, matched N.",
    "  H7  COMPOSITION vs SYNTAX. Motif identity and motif arrangement contribute separably; the",
    "      syntax arm (fixed motif set) is informative beyond the composition arms.",
    "",
    "EXPLORATORY - reported as such, may lack power",
    "  E1  Scaling EXPONENT differs by strategy (see Schedule for why this is not confirmatory).",
    "  E2  COMPOSITION / enrichment in the pooled arms: which reservoir does each acquisition draw",
    "      from? Enrichment = picked/expected; needs EQUAL numbers per reservoir in the union pool.",
    "      Preference is NOT value - descriptive; H1 licenses the informativeness claim.",
    "  E3  SUB-POOL arms (genomic-only / synthetic-only): does acquisition behaviour depend on what is",
    "      available to select from?",
    "  E4  MODEL-DEFINED vs BIOLOGY-DEFINED diversity: model-latent vs TFBS-embedding selection.",
    "      Disagreement between them is the interesting outcome.",
    "  E5  COMPOSITION-ONLY informativeness: uniform vs dinuc-shuffled vs composition-tuned random.",
    "  E6  A3 extrapolation beyond the measured range.",
    "",
    "PRE-SCREEN BEFORE ORDERING - free, uses the existing 300k model",
    "  Score the whole pool under every candidate acquisition; take pairwise rank correlation AND",
    "  Jaccard overlap of the top-N selections; cluster; fund one representative per cluster. The same",
    "  run returns the exact UNION size, which sets N so the order lands on budget.",
    "  Peter expects strong correlation within the uncertainty family and within the diversity family,",
    "  and uncertainty in regression tracks activity because larger values carry larger absolute error.",
    "  Activity-stratified is NOT a high-activity prior - it samples UNIFORMLY, the opposite of",
    "  uncertainty. Keep both, but confirm with the screen.",
]
for k, t in enumerate(H, start=2):
    c = ws2.cell(k, 1, t)
    c.font = Font(size=9, bold=(t.strip() != "" and not t.startswith("  ")))
ws2.column_dimensions["A"].width = 108


# ==================== MOTIF DESIGN ====================
wsmo = wb.create_sheet("MotifDesign")
wsmo["A1"] = "Motif reservoirs: vocabulary size is the design lever"
wsmo["A1"].font = Font(bold=True, size=13)
ML = [
    "THE PROBLEM",
    "  A 200 bp sequence holds ~4 motifs (each ~12 bp, plus spacing and flanks). The non-redundant human",
    "  vocabulary is ~600 motifs. So no sequence can contain 'the fixed set' - every sequence draws a",
    "  small SUBSET from a vocabulary. The design is therefore hierarchical, with four nested levels:",
    "     VOCABULARY V  - which motifs are eligible for the whole arm",
    "     SUBSET S      - which k motifs go into THIS sequence          <- composition",
    "     ARRANGEMENT   - order, spacing, orientation, copy number of S  <- syntax",
    "     BACKGROUND    - what S is embedded in",
    "",
    "WHY COMPOSITION AND SYNTAX CANNOT BOTH BE ESTIMATED FROM ONE POOL",
    "  If every sequence has a different subset AND a different arrangement, the two are perfectly",
    "  confounded at the sequence level. Separating them requires REPEATED MEASURES: the same subset",
    "  appearing in many arrangements, and the same arrangement realised with many subsets.",
    "  Varying background, motifs, spacing and order all at once across sequences - the intuitive design -",
    "  gives maximum coverage and ZERO ability to attribute. Both are needed, in different arms.",
    "",
    "THE ARITHMETIC THAT DECIDES IT  (arm = ~107k sequences, k = 4 slots)",
    "     |V|      per-motif obs      motif pairs in V      per-PAIR obs",
    "     600            498                179,700              2.5",
    "     300            996                 44,850             10.0",
    "     120          2,489                  7,140             62.7",
    "      60          4,978                  1,770            253.1",
    "  A FULL vocabulary gives excellent MARGINAL coverage (~500 observations per motif) and essentially",
    "  no PAIRWISE coverage (2.5 per pair). You cannot learn how motifs combine from a 600-motif pool at",
    "  this budget - there are 179,700 pairs and only ~450k pair-slots to spread over them.",
    "  A RESTRICTED vocabulary (~60) inverts this: ~253 observations per PAIR, which powers interaction",
    "  and spacing effects.",
    "",
    "SO THE ARMS SPLIT ON VOCABULARY SIZE, NOT ON 'COMPOSITION vs SYNTAX'",
    "  COVERAGE arms - FULL vocabulary, one random arrangement per random subset, background randomised.",
    "     Purpose: the model sees every motif many times in many contexts. This is what answers the",
    "     generalisation worry - and the vocabulary should be chosen to COVER every motif that occurs in",
    "     the eval sets (a free in-silico task: scan the eval batteries for motif content first).",
    "  SYNTAX CORE arm - RESTRICTED vocabulary (~60 motifs, chosen for eval-set relevance and TF-family",
    "     diversity), with subset x order x spacing x orientation x background deliberately CROSSED and",
    "     replicated. Purpose: attribution. This is the only arm that can decompose variance into",
    "     composition / syntax / background, because it is the only one with repeated measures.",
    "",
    "WITHIN THE SYNTAX CORE - a worked allocation for ~107k",
    "     ~400 motif subsets  x  4 orders  x  5 spacings  x  2 backgrounds  x  ~7 replicate draws",
    "  Every factor is crossed, so each main effect and the two-way interactions are estimable, and the",
    "  replicate draws give a pure-error term for the variance decomposition.",
    "",
    "SYNTAX AND CELL TYPE",
    "  Spacing and order requirements may well differ between lines - different TFs, different",
    "  cofactor geometry. You do NOT need a cell-type-specific syntax arm to see it: build the",
    "  syntax core with a vocabulary spanning both lines' enriched motifs, measure in both, and the",
    "  effect appears as a syntax x cell-type INTERACTION in the readout. Design-side specificity",
    "  would only be needed to TUNE syntax per line, which is a follow-up, not this study.",
    "",
    "CONSEQUENCE FOR H7",
    "  H7 (composition and syntax contribute separably) is tested INSIDE the syntax core via variance",
    "  decomposition, not by comparing two arms to each other. Comparing arms would confound the",
    "  vocabulary change with the design change.",
    "",
    "OPEN - CONFIRM BEFORE ORDERING",
    "  Exact non-redundant vocabulary size after clustering (JASPAR CORE / HOCOMOCO are ~700-1400 raw",
    "  models but collapse substantially). The ~600 figure above is illustrative; recompute it, since",
    "  every number on this sheet scales with it.",
    "  Motifs per sequence k: 4 assumed for 200 bp. If the assay length changes, k and all coverage",
    "  numbers change with it.",
]
for k_, t in enumerate(ML, start=2):
    c = wsmo.cell(k_, 1, t)
    c.font = Font(size=9, bold=(t.strip() != "" and not t.startswith("  ")))
wsmo.column_dimensions["A"].width = 104

# ==================== CELL-TYPE DESIGN ====================
wsct = wb.create_sheet("CellTypeDesign")
wsct["A1"] = "Joint / differential reservoirs, and why halves rather than separate arms"
wsct["A1"].font = Font(bold=True, size=13)
CL = [
    "HOW THE THREE CT TAGS ACTUALLY WORK",
    "  agn   - design uses no cell-type signal at all. One pool, assayed in both lines.",
    "  joint - design uses signal SHARED by both lines: regions open in BOTH, or motifs for TFs",
    "          expressed in BOTH. Tests whether shared regulatory information is sufficient.",
    "  diff  - design uses signal that DIFFERS between lines. The arm is built as a within-arm",
    "          factorial, not as a K562 half and a HepG2 half glued together:",
    "               K562-specific only  |  HepG2-specific only  |  BOTH in one sequence  |  neither",
    "          All four cells are assayed in both lines. The BOTH cell is what teaches differential",
    "          activity; the single cells give attribution; 'neither' is the internal control.",
    "",
    "WHY NOT TWO SEPARATE ARMS, ONE PER CELL TYPE - the reasons that are NOT about budget",
    "  1. EVERY SEQUENCE IS MEASURED IN BOTH LINES ANYWAY. The library is assayed in K562 and HepG2, so",
    "     a K562-specific-accessible sequence still returns a HepG2 measurement. Splitting into two arms",
    "     buys no extra information per sequence; it only changes which sequences exist.",
    "  2. THE TRAINING CORPUS IS SHARED. One two-output model trains on all of it. Separate arms would",
    "     only matter if separate models were trained on separate arms - which halves the data for each",
    "     and is strictly worse.",
    "  3. WITHIN-ARM CONTRASTS ARE PAIRED, BETWEEN-ARM CONTRASTS ARE NOT. The differential comparison",
    "     inside one arm is free of arm-level nuisance variance (synthesis batch, plate, sequencing",
    "     depth). Across two arms, that nuisance sits directly on top of the effect of interest.",
    "  4. THE QUESTION IS ABOUT THE DIFFERENTIAL, NOT ABOUT K562 IN ISOLATION. Carl: shared TFs (MYC,",
    "     AP1) mean most K562-active sequences are also HepG2-active. 'Does cell-type-specific",
    "     information help?' is answered by contrasting the differential design against the shared",
    "     design - not by running K562-specific and HepG2-specific separately, which under symmetry",
    "     estimates the same effect twice.",
    "  5. ASYMMETRY IS STILL TESTABLE. If HepG2 behaves worse (noisier assay, fewer characterised CREs),",
    "     the within-arm factorial exposes it: the K562-only and HepG2-only cells are compared directly.",
    "     At ~107k per arm each factorial cell holds ~27k sequences, which is enough to see it.",
    "     Two separate arms would spend 2x the budget to answer a question one arm already answers.",
    "",
    "THE ACQUISITION PER-CT ARM - N/2 + N/2 is the DECISION-RELEVANT question, not a compromise",
    "  Matched N is the definition of a controlled comparison: at 2N you cannot tell whether per-CT",
    "  selection won because it selects better or because it had twice the data. On a scaling curve that",
    "  ambiguity gets worse, since more data always helps.",
    "  But the deeper point is that N/2 + N/2 IS the real-world question. You are ordering ONE library",
    "  with a FIXED budget. The decision facing you is 'should I spend my budget selecting jointly, or",
    "  split it and select per cell type?' - which is exactly what the matched-N arm answers.",
    "  'Run per-CT selection at full scale in each line' is a different question (it doubles the order),",
    "  and it is not the one a fixed library budget poses.",
    "",
    "NESTING GIVES THE OTHER COMPARISONS FOR FREE",
    "  Because every arm is trained on nested subsets, the joint arm can be subsampled to N/2 after the",
    "  fact. So from the same ordered sequences you can read off:",
    "     joint at N        vs  per-CT at N/2+N/2     - budget-matched (the decision question)",
    "     joint at N/2      vs  K562-only at N/2      - per-cell-type-matched (the mechanism question)",
    "  No extra sequences are needed for either. This is the main reason the halving costs nothing",
    "  analytically - the comparison you gave up is recoverable from the scaling curve.",
]
for k_, t in enumerate(CL, start=2):
    c = wsct.cell(k_, 1, t)
    c.font = Font(size=9, bold=(t.strip() != "" and not t.startswith("  ")))
wsct.column_dimensions["A"].width = 104


# ==================== STRATA ====================
wsst = wb.create_sheet("Strata")
wsst["A1"] = "What the CT tags mean, and what lives inside each arm"
wsst["A1"].font = Font(bold=True, size=13)
SL = [
    "THE TAG DESCRIBES THE DESIGN INPUT, NOT THE RESULT",
    "  Every arm is assayed in BOTH cell lines, so every arm returns a K562 measurement and a HepG2",
    "  measurement. Cell-type-specific EFFECTS are therefore measurable in every single arm, including",
    "  the agnostic ones. The tag answers only one question:",
    "        does the DESIGN STEP have to consult cell-type-specific information?",
    "     agn   - no. Uniform random, dinuc-shuffle, EP-PCR, full-vocabulary motifs, HyenaDNA.",
    "     joint - yes, but only information SHARED by both lines.",
    "     diff  - yes, information that DIFFERS between the lines.",
    "  This is why 'syntax is potentially cell-type specific' does NOT require a cell-type-specific",
    "  syntax arm: a syntax x cell-type interaction is recovered from the READOUT of the agnostic",
    "  syntax core, because the same sequences are measured in both lines. Design-side specificity is",
    "  only needed if you want to TUNE the syntax separately per line, which is a later question.",
    "",
    "THE DUPLICATION IS FIXED",
    "  Previously 'ENCODE open in BOTH' was its own arm AND a cell inside the differential arm - the",
    "  same sequences counted twice. Likewise 'shared-core motifs' duplicated a stratum of the",
    "  CT-vocabulary arm. Both are now single arms with internal strata. That removed 2 arms and pushed",
    "  every arm from ~107k to ~114k.",
    "",
    "ENCODE ACCESSIBILITY - one arm, a PARTITION of the genome (not a constructed factorial)",
    "  Each candidate region either is or is not accessible in each line, so every region falls in",
    "  exactly one stratum. Nothing is constructed and nothing overlaps:",
    "     open in BOTH          - tests whether SHARED accessibility is sufficient",
    "     K562-accessible only  - differential, direction 1",
    "     HepG2-accessible only - differential, direction 2",
    "     open in NEITHER       - real-sequence negatives (absorbs the 'closed regions' proposal)",
    "  All four are assayed in both lines, so the K562-only stratum still yields HepG2 labels. That is",
    "  what makes the specificity comparison bidirectional.",
    "",
    "MOTIF CT-VOCABULARY - one arm, a CONSTRUCTED 2x2 (this one you do build)",
    "  Here you choose what to insert, so the strata are defined by which motifs are placed:",
    "     shared-core motifs only     - the MYC / AP1 core Carl described",
    "     K562-enriched motifs only   - differential, direction 1",
    "     HepG2-enriched motifs only  - differential, direction 2",
    "     BOTH in the same sequence   - the cell that teaches differential activity",
    "  Attribution survives because the single-vocabulary strata sit in the same arm as the mixed one.",
    "",
    "ARE THE STRATA UNDERPOWERED? - the arithmetic",
    "  At ~114k per arm, a 4-way partition gives ~28.6k per stratum. For scale: our entire current",
    "  bake-off runs at D = 30k, so each stratum is about the size of a complete existing experiment.",
    "  A stratum supports a nested curve from 1k to ~28.6k, roughly 1.5 OOMs - shorter than an arm's",
    "  ~2 OOMs, but real. It also supports the within-arm contrasts, which are PAIRED and therefore",
    "  higher-powered than the equivalent between-arm comparison.",
    "  What a stratum does NOT support is the full 2-OOM curve. So the decision to put to the PIs is",
    "  exactly this:",
    "     STRATUM  - cheap, gives a paired contrast plus a ~1.5-OOM curve         (current default)",
    "     ARM      - costs a full ~114k and shrinks every other arm, gives ~2 OOMs and a clean",
    "                between-arm comparison",
    "  Promote a stratum to an arm only where the per-stratum SCALING RATE is itself the hypothesis.",
    "",
    "WHY STRATA ARE THE RIGHT DEFAULT HERE",
    "  The reservoir-level claim ('accessibility-based selection is informative') uses the whole arm,",
    "  all ~114k. Only the specificity claim needs the strata, and that claim is a CONTRAST between",
    "  strata - which is precisely the comparison that gains from being inside one arm, since",
    "  arm-level nuisance variance (synthesis batch, plate, sequencing depth) cancels.",
]
for k_, t in enumerate(SL, start=2):
    c = wsst.cell(k_, 1, t)
    c.font = Font(size=9, bold=(t.strip() != "" and not t.startswith("  ")))
wsst.column_dimensions["A"].width = 104

wb.save(OUT)
w_in = (WA + WB + WC + WD * nA) * 7 / 96
h_in = (SROW - AH + 1) * 14.0 / 72
print("wrote", OUT)
print(
    f"  PASTE BLOCK A{AH}:{LC}{SROW} = {SROW - AH + 1} rows x {nA + 3} cols, "
    f"{w_in:.2f} x {h_in:.2f} in  (16:9 usable ~9.5 x 4.8)"
)
print(f"  toggles {KC0}{A0}:{KCL}{ALAST}   arms = {len(ON)} -> {3_200_000 / len(ON):,.0f} seq/arm")
print("  sheets:", wb.sheetnames)
