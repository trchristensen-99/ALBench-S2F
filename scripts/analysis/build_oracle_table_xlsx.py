"""Editable oracle-accuracy table (.xlsx): individual model out-of-fold vs the 10-fold ensemble.

Two quantities per evaluation set, each reported as both r and MSE:
  individual, out-of-fold   each sequence scored only by the fold that held it out - one model,
                            honestly evaluated.
  ensemble of 10            the 10-fold mean prediction, i.e. the deployed oracle. Note every
                            sequence was in 9 of the 10 folds' training data, so this is not a
                            held-out number.

r and MSE answer different questions and can disagree: r is scale-free and says whether the
ranking is right, MSE says whether the values are right. That distinction matters here because the
predictions are used as training targets, where absolute calibration counts.

All cells are real values, so font, column width and number format stay editable, and the block
pastes into PowerPoint or Slides as a native table.
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HDR = PatternFill("solid", fgColor="1E293B")
BAND = PatternFill("solid", fgColor="F4F6F9")
thin = Side(style="thin", color="C8CFD8")

# (label, n, oof_r, oof_mse, ens_r, ens_mse)
ROWS = [
    ("Genomic reference", 30659, 0.9496, 0.1310, 0.9745, 0.0669),
    ("Designed high-activity", 22962, 0.8398, 0.7569, 0.9815, 0.0994),
    ("Negative controls", 471, 0.8231, 0.0790, 0.9274, 0.0340),
    ("SNV alleles (absolute)", 56144, 0.9224, 0.1653, 0.9569, 0.0935),
    ("SNV effect (alt - ref)", 29493, 0.3928, 0.1861, 0.4621, 0.1728),
]
COLS = [
    "evaluation set",
    "n",
    "single model\nout-of-fold: r",
    "single model\nout-of-fold: MSE",
    "ensemble\nof 10: r",
    "ensemble\nof 10: MSE",
]
WIDTH = [30, 10, 14, 15, 13, 14]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--out", default=os.path.expanduser("~/Downloads/notion_updates/oracle_label_quality.xlsx")
    )
    ap.add_argument("--decimals", type=int, default=3)
    a = ap.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "Oracle label quality"
    fmt = "0." + "0" * a.decimals

    ws["A1"] = "Oracle label quality"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Pearson r against measured K562 activity"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    HROW = 4
    for j, c in enumerate(COLS, start=1):
        cell = ws.cell(HROW, j, c)
        cell.fill = HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(
            horizontal="left" if j == 1 else "right", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=thin)
    ws.row_dimensions[HROW].height = 30

    for i, (label, n, o_r, o_m, e_r, e_m) in enumerate(ROWS):
        r = HROW + 1 + i
        vals = [label, n, o_r, o_m, e_r, e_m]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(r, j, v)
            cell.font = Font(size=10, bold=(j in (3, 4)))
            if i % 2 == 1:
                cell.fill = BAND
            if j == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0" if j == 2 else fmt
        ws.cell(r, 1).border = Border(top=thin) if i == len(ROWS) - 1 else Border()

    # derived contrasts, as live formulas so edits propagate
    d = HROW + len(ROWS) + 2
    ws.cell(d, 1, "Ensemble advantage over one held-out model").font = Font(bold=True, size=10)
    for i in range(len(ROWS)):
        src = HROW + 1 + i
        c1 = ws.cell(d + 1, 3 + 0, f"=E{src}-C{src}")
        c2 = ws.cell(d + 1, 3 + 1, f"=F{src}-D{src}")
    ws.cell(d + 1, 1, "delta r  /  delta MSE  (per row, see columns)").font = Font(
        size=9, italic=True, color="666666"
    )
    for i in range(len(ROWS)):
        src = HROW + 1 + i
        cr = ws.cell(d + 2 + i, 3, f"=E{src}-C{src}")
        cm = ws.cell(d + 2 + i, 4, f"=F{src}-D{src}")
        ws.cell(d + 2 + i, 1, ROWS[i][0]).font = Font(size=9.5)
        for c in (cr, cm):
            c.number_format = "+0." + "0" * a.decimals + ";-0." + "0" * a.decimals
            c.font = Font(size=9.5)
            c.alignment = Alignment(horizontal="right")

    note = (
        "single model out-of-fold = each sequence scored only by the fold that held it out.\n"
        "ensemble of 10 = the 10-fold mean prediction (the deployed oracle). Every sequence was in "
        "9 of the 10 folds' training data, so this column is not held out.\n"
        "MSE is on log2FC units, the same scale as the labels. Lower is better; r and MSE can "
        "disagree, since r ignores calibration.\n"
        "SNV rows use true single-nucleotide substitutions only. SNV alleles pools ref and alt. "
        "The effect row requires both alleles scored by the same fold, and is genome-wide."
    )
    nr = d + 3 + len(ROWS)
    ws.cell(nr, 1, note).font = Font(size=8.5, color="555555")
    ws.cell(nr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr + 4, end_column=6)

    for j, w in enumerate(WIDTH, start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    os.makedirs(os.path.dirname(a.out), exist_ok=True)
    wb.save(a.out)
    print(f"wrote {a.out}")
    print(f"  paste A{HROW}:F{HROW + len(ROWS)} into Slides or PowerPoint as a native table")


if __name__ == "__main__":
    main()
