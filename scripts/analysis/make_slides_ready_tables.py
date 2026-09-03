"""Two routes to WELL-FORMATTED, EDITABLE tables in Google Slides.

Route 1 (instant): an .xlsx whose column widths, wrapping and header styling are already set. Column
widths carry through Sheets -> Slides on paste, so the pasted table arrives correctly proportioned
instead of needing manual resizing. Result is a NATIVE Slides table: fully editable mid-presentation.

Route 2 (repeatable): a Google Apps Script that builds the tables directly in the deck with exact
column widths, font sizes and status colours. Run once, get every table as a native editable table.
"""

import json
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = os.path.expanduser("~/Downloads/pi_meeting_figs/editable_tables")
OUT_X = os.path.join(SRC, "SLIDES_READY.xlsx")

# (sheet name, csv, per-column width in characters)
SPECS = [
    ("WhatVaries", "what_varies_matrix.csv", [26, 20, 22, 20, 22]),
    ("Reservoir_1", None, [18, 30, 26, 26, 13]),
    ("Reservoir_2", None, [18, 30, 26, 26, 13]),
    ("Acquisition", "acquisition.csv", [20, 44, 18, 26]),
    ("EvalSets", "eval_sets.csv", [22, 34, 8, 20, 26]),
    ("HP_axes_summary", "hp_axes.csv", [26, 32, 22, 30]),
    ("HP_1_architecture", "HP_1_architecture.csv", [24, 14, 26, 34, 16]),
    ("HP_2_optimisation", "HP_2_optimisation.csv", [24, 14, 24, 36, 16]),
    ("HP_3_FM_specific", "HP_3_FM_specific.csv", [28, 14, 30, 32, 14]),
    ("HP_4_strategies", "HP_4_strategies.csv", [26, 36, 40]),
]
res = pd.read_csv(f"{SRC}/reservoir_FULL.csv")[
    [
        "Family",
        "Motif source (how identified)",
        "Background (how built)",
        "Placement rule",
        "Status",
    ]
]
frames = {"Reservoir_1": res.iloc[:9], "Reservoir_2": res.iloc[9:]}

STATUS_FILL = {
    "KEEP": "DCFCE7",
    "ADD": "DBEAFE",
    "ADD (JB)": "C7D2FE",
    "REVISE": "FEF3C7",
    "PROPOSED": "E9D5FF",
    "PARAMETERISE": "FED7AA",
    "REMOVE": "FECACA",
    "DEPRIORITISE": "E5E7EB",
    "IMPLEMENTED": "DCFCE7",
    "NOT IMPLEMENTED": "FECACA",
}

with pd.ExcelWriter(OUT_X, engine="openpyxl") as xl:
    for sheet, csv, widths in SPECS:
        df = frames[sheet] if csv is None else pd.read_csv(f"{SRC}/{csv}")
        df.to_excel(xl, sheet_name=sheet, index=False)
        ws = xl.sheets[sheet]
        for i, w in enumerate(widths[: len(df.columns)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in range(1, len(df.columns) + 1):  # header
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        status_col = next((i + 1 for i, c in enumerate(df.columns) if c.lower() == "status"), None)
        for r in range(2, len(df) + 2):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=8)
            if status_col:
                v = str(ws.cell(row=r, column=status_col).value).strip()
                if v in STATUS_FILL:
                    ws.cell(row=r, column=status_col).fill = PatternFill(
                        "solid", fgColor=STATUS_FILL[v]
                    )
        ws.freeze_panes = "A2"
print(f"wrote {OUT_X}")

# ---- Route 2: Apps Script with the data inlined
payload = {}
for sheet, csv, widths in SPECS:
    df = frames[sheet] if csv is None else pd.read_csv(f"{SRC}/{csv}")
    payload[sheet] = {
        "cols": list(df.columns),
        "rows": df.fillna("").astype(str).values.tolist(),
        "widths": widths[: len(df.columns)],
    }

gs = """/**
 * Build well-formatted, EDITABLE tables in this Google Slides deck.
 *
 * HOW TO RUN
 *   1. Open your deck -> Extensions -> Apps Script
 *   2. Delete anything there, paste this whole file, click Save
 *   3. Select buildAllTables in the function dropdown -> Run
 *      (first run asks for permission to edit the presentation)
 *
 * Each table is inserted on its own new slide as a NATIVE Slides table, so it stays fully editable
 * during the presentation. Column widths, font sizes and status colours are set for you, which is
 * the part that is painful to do by hand.
 */
var TABLES = %s;

var STATUS_COLORS = {
  'KEEP': '#DCFCE7', 'ADD': '#DBEAFE', 'ADD (JB)': '#C7D2FE', 'REVISE': '#FEF3C7',
  'PROPOSED': '#E9D5FF', 'PARAMETERISE': '#FED7AA', 'REMOVE': '#FECACA',
  'DEPRIORITISE': '#E5E7EB', 'IMPLEMENTED': '#DCFCE7', 'NOT IMPLEMENTED': '#FECACA'
};

function buildAllTables() {
  var deck = SlidesApp.getActivePresentation();
  Object.keys(TABLES).forEach(function (name) { buildTable(deck, name, TABLES[name]); });
}

function buildTable(deck, name, spec) {
  var slide = deck.appendSlide(SlidesApp.PredefinedLayout.BLANK);
  slide.insertTextBox(name, 20, 8, 400, 24).getText().getTextStyle().setFontSize(14).setBold(true);

  var nRows = spec.rows.length + 1, nCols = spec.cols.length;
  var table = slide.insertTable(nRows, nCols, 20, 40, 680, Math.min(420, 22 * nRows));

  var total = spec.widths.reduce(function (a, b) { return a + b; }, 0);
  for (var c = 0; c < nCols; c++) {
    table.getColumn(c).setWidth(680 * spec.widths[c] / total);   // proportional, set once
  }
  for (var c = 0; c < nCols; c++) {
    var h = table.getCell(0, c);
    h.getText().setText(spec.cols[c]);
    h.getText().getTextStyle().setFontSize(8).setBold(true).setForegroundColor('#FFFFFF');
    h.getFill().setSolidFill('#1E293B');
  }
  var statusCol = spec.cols.map(function (x) { return x.toLowerCase(); }).indexOf('status');
  for (var r = 0; r < spec.rows.length; r++) {
    for (var c = 0; c < nCols; c++) {
      var cell = table.getCell(r + 1, c);
      cell.getText().setText(spec.rows[r][c]);
      cell.getText().getTextStyle().setFontSize(7);
      if (c === statusCol && STATUS_COLORS[spec.rows[r][c]]) {
        cell.getFill().setSolidFill(STATUS_COLORS[spec.rows[r][c]]);
      } else if (r %% 2 === 1) {
        cell.getFill().setSolidFill('#F8FAFC');
      }
    }
  }
}
""" % json.dumps(payload, indent=1)

gs_path = os.path.join(SRC, "BuildSlidesTables.gs")
open(gs_path, "w").write(gs)
print(f"wrote {gs_path}")
